from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from Constants import globalConstants
# import globalConstants
#prefix => ön ek test_
# postfix
#Sadece test fonksiyonları çalıştırmak zorunda olmadığından;
# test fonksiyonları test_ ile başlar.

class Test_DemoClass:
    # Her testten önce çağrılır
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath= str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True) 
        # exist_ok=True olması, halihazırda bir klasör var mı varsa oluşturmadan devam et demektir
        # 25.07.2023
        
        # günün tarihini al bu tarih ile klasör var mı kontrol et yoksa oluştur
    # Her testten sonra çağrılır    
    def teardown_method(self):
        self.driver.quit()
        
    # setup -> test_demofunc -> teardown
    def test_demoFunc(self):
        # 3A Act Arrange Assert
        text = "Hello"
        assert text == "Hello"
        # setup -> test_demo2 -> teardown
    def test_demo2(self):
            # assert, testi sonuçlandırmaya yarar.
            # assert, ilgili testin sonucunu br koşula bağlıyor
            assert True

    # decoratorlerden çağırılan fonksiyona self parametresi verilmez
    def getData():
         # veriyi al
         #bir excel dosyasını eklemee yarayan fonksiyon
         excelfile = openpyxl.load_workbook("data/invalid_login.xlsx")
         SelectedSheet = excelfile["Sheet1"]

         totalRows = SelectedSheet.max_row
         data = []
         for i in range(2,totalRows+1):
                username = SelectedSheet.cell(i,1).value
                password = SelectedSheet.cell(i,2).value
                toppleData = (username,password)
                data.append(toppleData)

         # exceldeki hücreyi bulmak için cell() fonksiyonu kullanılır .value hücre içindeki değeri alır

         return data
    
        # decorators
    # @pytest.mark.skip() #sonraki fonkiyonu çalıştırmama komutu
    @pytest.mark.parametrize("username, password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementvisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementvisible((By.ID,"password"),10)
        passwordInput =self.driver.find_element(By.ID, "password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID, "login-button")  
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
    def waitForElementvisible(self,locator,timeout=5):
         WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
         
















        